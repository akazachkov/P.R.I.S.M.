# app/core/elements/convert_register_to_sum_sn.py

import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Callable, List
from collections import Counter, defaultdict, deque
import pandas as pd
from openpyxl import load_workbook

# ----------------------------------------------------------------------
# Патч для openpyxl (игнорирование extLst)
_openpyxl_patched = False


def apply_openpyxl_patch():
    global _openpyxl_patched
    if _openpyxl_patched:
        return
    try:
        import openpyxl.styles.fills

        @classmethod
        def patched_from_tree(cls, node):
            attrib = dict(node.attrib) if hasattr(node, 'attrib') else {}
            if 'extLst' in attrib:
                del attrib['extLst']
            return cls(**attrib)

        openpyxl.styles.fills.PatternFill._from_tree = patched_from_tree
        _openpyxl_patched = True
    except Exception:
        pass


def safe_read_excel(file_path, **kwargs):
    try:
        return pd.read_excel(file_path, **kwargs)
    except TypeError as e:
        if 'extLst' in str(e):
            apply_openpyxl_patch()
            return pd.read_excel(file_path, **kwargs)
        else:
            raise


def safe_load_workbook(file_path, **kwargs):
    try:
        return load_workbook(file_path, **kwargs)
    except TypeError as e:
        if 'extLst' in str(e):
            apply_openpyxl_patch()
            return load_workbook(file_path, **kwargs)
        else:
            raise


# ----------------------------------------------------------------------
# Вспомогательные функции
def natural_sort_key(serial):
    if not serial:
        return (0, '', '')
    parts = re.split(r'(\d+)', serial)
    key_parts = []
    for part in parts:
        if part:
            if part.isdigit():
                key_parts.append((0, int(part)))
            else:
                key_parts.append((1, part.lower()))
    return tuple(key_parts)


def format_cell_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    elif isinstance(value, (int, float)):
        try:
            if 0 < value < 100000:
                excel_epoch = datetime(1899, 12, 30)
                date_value = excel_epoch + pd.Timedelta(days=value)
                return date_value.strftime("%d.%m.%Y")
        except (ValueError, TypeError, OverflowError):
            pass
    return str(value).strip()


def get_additional_data(file_path, log_func):
    try:
        wb = safe_load_workbook(file_path, data_only=True)
        ws = wb.active
        i5 = format_cell_value(ws['I5'].value)
        i9 = format_cell_value(ws['I9'].value)
        e12 = format_cell_value(ws['E12'].value)
        ak14 = format_cell_value(ws['AK14'].value)
        ak15 = format_cell_value(ws['AK15'].value)
        g22 = format_cell_value(ws['G22'].value)
        combined_ak = f"'{ak14}' от '{ak15}'" if ak14 or ak15 else ''
        return [i5, i9, e12, combined_ak, g22]
    except Exception as e:
        log_func(f"Ошибка извлечения доп. данных: {e}")
        return ['', '', '', '', '']


def check_numbering_order(df, name_col, log_func):
    try:
        number_col = None
        possible_number_cols = ['№ п/п']
        for col in df.columns:
            col_clean = str(col).strip().lower()
            if any(p.lower() in col_clean for p in possible_number_cols):
                number_col = col
                break
        if not number_col:
            return None
        numbers = []
        for idx, row in df.iterrows():
            if pd.isna(row[number_col]):
                continue
            if pd.isna(row[name_col]) or str(row[name_col]).strip() == '':
                continue
            try:
                num_value = str(row[number_col]).strip()
                if num_value:
                    num_clean = re.sub(r'[^\d]', '', num_value.split('.')[0])
                    if num_clean:
                        num_int = int(num_clean)
                        numbers.append(num_int)
            except (ValueError, TypeError):
                continue
        if not numbers:
            return None
        numbers_sorted = sorted(numbers)
        expected = list(range(min(numbers_sorted), max(numbers_sorted) + 1))
        has_gaps = numbers_sorted != expected
        has_duplicates = len(numbers) != len(set(numbers))
        starts_from_one = min(numbers_sorted) == 1
        return (not has_gaps) and (not has_duplicates) and starts_from_one
    except Exception as e:
        log_func(f"Ошибка проверки нумерации: {e}")
        return None


# ----------------------------------------------------------------------
# Основная функция обработки ведомости (ВСО)
def process_file(input_file_path, log_func):
    try:
        header_row = -1
        target_headers = [
            'Наименование оборудования',
            'Серийный номер оборудования',
            'Количество, шт'
        ]
        df_preview = safe_read_excel(input_file_path, header=None, nrows=50)
        for row_idx in range(len(df_preview)):
            row_cells = [
                str(cell).strip()
                if not pd.isna(cell)
                else '' for cell in df_preview.iloc[row_idx]
            ]
            if sum(target in row_cells for target in target_headers) == 3:
                header_row = row_idx
                break
        if header_row == -1:
            df = safe_read_excel(input_file_path)
        else:
            df = safe_read_excel(input_file_path, header=header_row)
        df.columns = [str(col).strip() for col in df.columns]
        name_col = serial_col = qty_col = None
        for col in df.columns:
            col_lower = col.lower()
            if all(keyword in col_lower for keyword in (
                'наименование', 'оборудования'
            )):
                name_col = col
            elif all(keyword in col_lower for keyword in (
                'серийный', 'номер', 'оборудования'
            )):
                serial_col = col
            elif all(keyword in col_lower for keyword in ('количество', 'шт')):
                qty_col = col
        if not all([name_col, serial_col, qty_col]):
            raise Exception(
                f"Необходимые столбцы не найдены в файле {input_file_path}!"
            )
        numbering_check = check_numbering_order(df, name_col, log_func)
        additional_data = get_additional_data(input_file_path, log_func)
        items_dict = {}
        for idx, row in df.iterrows():
            try:
                item_name = (
                    '' if pd.isna(row[name_col])
                    else str(row[name_col]).strip()
                )
                serial_numbers = (
                    '' if pd.isna(row[serial_col])
                    else str(row[serial_col]).strip()
                )
                qty = row[qty_col] if not pd.isna(row[qty_col]) else 0
                if pd.isna(item_name) or item_name == '':
                    continue
                try:
                    qty_str = str(qty).replace(',', '.')
                    qty_clean = re.sub(r'[^\d.]', '', qty_str)
                    qty_int = int(float(qty_clean)) if qty_clean else 0
                except (ValueError, TypeError):
                    continue
                if qty_int <= 0:
                    continue
                serial_list = []
                if serial_numbers and serial_numbers.lower() not in [
                        '', 'nan', 'none']:
                    serial_str = re.sub(r'[\n\r\t]+', '|', serial_numbers)
                    serial_str = re.sub(r'[,;]+', '|', serial_str)
                    serial_str = re.sub(r'\s+', '|', serial_str)
                    for part in serial_str.split('|'):
                        part_clean = part.strip()
                        if part_clean and part_clean not in ['-', '—', '–']:
                            serial_list.append(part_clean)
                serial_list.sort(key=natural_sort_key)
                if item_name not in items_dict:
                    items_dict[item_name] = {
                        'serial_numbers': [], 'quantity': 0
                    }
                items_dict[item_name]['serial_numbers'].extend(serial_list)
                items_dict[item_name]['quantity'] += qty_int
            except Exception:
                continue
        return (
            items_dict, name_col, serial_col, numbering_check, additional_data
        )
    except Exception as e:
        raise Exception(f"Ошибка чтения файла {input_file_path}: {e}")


# ----------------------------------------------------------------------
# Функции для ОС-15 (поддержка нескольких файлов)
def find_oc15_files(oc15_folder: Path, normalized_id: str) -> List[Path]:
    """
    Возвращает список всех файлов ОС-15 в указанной папке,
    имя которых начинается с числа, соответствующего normalized_id.
    """
    if not oc15_folder.exists():
        return []
    id_raw = str(int(normalized_id))  # убираем ведущие нули
    result = []
    for file_path in oc15_folder.glob("*.xlsx"):
        # Берём первую часть имени до пробела (если есть)
        first_part = file_path.stem.split()[0] if file_path.stem else ""
        digits = re.sub(r'\D', '', first_part)
        if digits and int(digits) == int(id_raw):
            result.append(file_path)
    return result


def process_oc15_file(
    file_path: Path, log_func: Optional[Callable] = None
) -> dict:
    """
    Обрабатывает один файл ОС-15, извлекает наименования и серийные номера.
    Возвращает items_dict.
    """
    if log_func is None:
        def log_func(msg, level='info'):
            pass
    try:
        wb = safe_load_workbook(file_path, data_only=True)
        ws = wb.active
        items_dict = {}
        current_name = None
        start_row = 29
        max_row = ws.max_row
        for row_idx in range(start_row, max_row + 1):
            name_cell = ws.cell(row=row_idx, column=1)
            serial_cell = ws.cell(row=row_idx, column=6)
            name_value = format_cell_value(name_cell.value)
            serial_value = format_cell_value(serial_cell.value)
            if not name_value and not serial_value:
                break
            if name_value:
                current_name = name_value
            if not current_name:
                continue
            serials = []
            if (
                serial_value
                and serial_value.lower()
                not in ['', 'nan', 'none']
            ):
                serial_str = re.sub(r'[\n\r\t]+', '|', serial_value)
                serial_str = re.sub(r'[,;]+', '|', serial_str)
                serial_str = re.sub(r'\s+', '|', serial_str)
                for part in serial_str.split('|'):
                    part_clean = part.strip()
                    if part_clean and part_clean not in ['-', '—', '–']:
                        serials.append(part_clean)
            if current_name not in items_dict:
                items_dict[current_name] = {
                    'serial_numbers': [], 'quantity': 0
                }
            items_dict[current_name]['serial_numbers'].extend(serials)
        wb.close()
        log_func(
            f"Обработан файл ОС-15: {len(items_dict)} наименований", "info"
        )
        return items_dict
    except Exception as e:
        log_func(f"Ошибка обработки файла ОС-15 {file_path}: {e}", "error")
        return {}


def process_oc15_files(
    file_paths: List[Path], log_func: Optional[Callable] = None
) -> dict:
    """
    Обрабатывает список файлов ОС-15, объединяя результаты.
    Возвращает items_dict с суммированными серийными номерами.
    """
    if log_func is None:
        def log_func(msg, level='info'):
            pass
    combined_dict = {}
    for file_path in file_paths:
        log_func(f"Обработка файла ОС-15: {file_path.name}", "info")
        items_dict = process_oc15_file(file_path, log_func)
        for name, data in items_dict.items():
            if name not in combined_dict:
                combined_dict[name] = {'serial_numbers': [], 'quantity': 0}
            combined_dict[name]['serial_numbers'].extend(
                data['serial_numbers']
            )
            combined_dict[name]['quantity'] += data['quantity']
    return combined_dict


# ----------------------------------------------------------------------
# Функция выравнивания с логгированием несовпадений
def align_items_dicts(items_dict_left, items_dict_right, log_func=None):
    """
    Выравнивает два словаря только по серийным номерам.
    Для номеров из ОС-15 дополнительно пробует изменить регистр и искать
    совпадение.
    """

    # Разворачиваем словари в списки (name, serial)
    left_items = [(name, serial) for name, data in items_dict_left.items()
                  for serial in data.get('serial_numbers', []) if serial]
    right_items = [(name, serial) for name, data in items_dict_right.items()
                   for serial in data.get('serial_numbers', []) if serial]

    left_serials = [s for _, s in left_items]
    right_serials = [s for _, s in right_items]
    count_left = Counter(left_serials)
    count_right = Counter(right_serials)

    # 1. Точные совпадения
    common_pairs = []
    temp_left = count_left.copy()
    temp_right = count_right.copy()
    for serial in sorted(
        set(left_serials) & set(right_serials), key=natural_sort_key
    ):
        k = min(temp_left[serial], temp_right[serial])
        for _ in range(k):
            common_pairs.append((serial, serial))
            temp_left[serial] -= 1
            temp_right[serial] -= 1

    # 2. Оставшиеся уникальные
    unique_left = []
    for serial, cnt in temp_left.items():
        if cnt > 0:
            unique_left.extend([serial] * cnt)
    unique_right = []
    for serial, cnt in temp_right.items():
        if cnt > 0:
            unique_right.extend([serial] * cnt)

    # 3. Дополнительные совпадения (смена регистра для правых)
    def swap_case(s):
        return ''.join(
            ch.lower() if ch.isupper()
            else ch.upper()
            if ch.islower()
            else ch for ch in s
        )

    left_unique_counter = Counter(unique_left)
    extra_pairs = []
    new_unique_right = []
    for serial in unique_right:
        swapped = swap_case(serial)
        if swapped != serial and left_unique_counter.get(swapped, 0) > 0:
            extra_pairs.append((swapped, serial))
            left_unique_counter[swapped] -= 1
        else:
            new_unique_right.append(serial)
    unique_left = []
    for serial, cnt in left_unique_counter.items():
        if cnt > 0:
            unique_left.extend([serial] * cnt)
    unique_right = new_unique_right

    # Логирование
    if log_func:
        if unique_left:
            log_func(f"Только в ВСО: {unique_left}", "warning")
        if unique_right:
            log_func(f"Только в ОС-15: {unique_right}", "warning")
        if extra_pairs:
            log_func(f"Совпали после смены регистра: {extra_pairs}", "info")

    # Очереди наименований
    left_queue = defaultdict(deque)
    for name, serial in left_items:
        left_queue[serial].append(name)
    right_queue = defaultdict(deque)
    for name, serial in right_items:
        right_queue[serial].append(name)

    left_rows, right_rows = [], []

    # Точные совпадения
    for left_serial, right_serial in common_pairs:
        left_rows.append((
            left_queue[left_serial].popleft()
            if left_queue[left_serial] else '', left_serial
        ))
        right_rows.append((
            right_queue[right_serial].popleft()
            if right_queue[right_serial] else '', right_serial
        ))

    # Дополнительные совпадения
    for left_serial, right_serial in extra_pairs:
        left_rows.append((
            left_queue[left_serial].popleft()
            if left_queue[left_serial] else '', left_serial
        ))
        right_rows.append((
            right_queue[right_serial].popleft()
            if right_queue[right_serial] else '', right_serial
        ))

    # Уникальные левые
    for serial in unique_left:
        for _ in range(
            count_left[serial] - min(
                count_left[serial], count_right.get(serial, 0)
            )
        ):
            left_rows.append((
                left_queue[serial].popleft()
                if left_queue[serial] else '', serial
            ))
            right_rows.append(('', ''))

    # Уникальные правые
    for serial in unique_right:
        for _ in range(
            count_right[serial] - min(
                count_right[serial], count_left.get(serial, 0)
            )
        ):
            left_rows.append(('', ''))
            right_rows.append((
                right_queue[serial].popleft()
                if right_queue[serial] else '', serial
            ))

    # Выравнивание
    max_len = max(len(left_rows), len(right_rows))
    left_rows += [('', '')] * (max_len - len(left_rows))
    right_rows += [('', '')] * (max_len - len(right_rows))

    return left_rows, right_rows


# ----------------------------------------------------------------------
# Создание итогового файла сверки
def create_combined_verification_file(
    items_dict_vso: dict,
    items_dict_oc15: dict,
    output_path: Path,
    log_func: Optional[Callable] = None,
    rma_dict: Optional[dict] = None,
    uploading_file_path: Optional[Path] = None
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    left_rows, right_rows = align_items_dicts(
        items_dict_vso, items_dict_oc15, log_func
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Сверка ВСО <--> ОС-15"

    # Заголовки (левая часть: A-E, разделитель F, правая часть: G-J)
    if right_rows:
        headers = [
            "№", "Наименование оборудования", "Серийный номер", "RMA",
            "Выгрузка",
            "",  # Разделитель
            "№", "Наименование оборудования", "Серийный номер", "RMA"
        ]
        ws.append(headers)
    else:
        headers = [
            "№", "Наименование оборудования", "Серийный номер", "RMA",
            "Выгрузка"
        ]
        ws.append(headers)

    bold_font = Font(bold=True)
    # Левые заголовки (столбцы A-E)
    for col in [1, 2, 3, 4, 5]:
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')
    if right_rows:
        # Правые заголовки (столбцы G-J)
        for col in [7, 8, 9, 10]:
            cell = ws.cell(row=1, column=col)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['F'].width = 3  # Разделитель

    # Заполнение данными
    for i, (left_row, right_row) in enumerate(
        zip(left_rows, right_rows), start=1
    ):
        left_name, left_serial = left_row
        row_num = i + 1
        # Левый блок
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=left_name)
        ws.cell(row=row_num, column=3, value=left_serial)

        # RMA левый (столбец D)
        if rma_dict and left_serial:
            rma_entry = rma_dict.get(left_serial)
            if rma_entry:
                val, headers = rma_entry
                cell_text = f"{val} ({headers})" if headers else val
                ws.cell(row=row_num, column=4, value=cell_text)
            else:
                ws.cell(row=row_num, column=4, value="")
        # Статус загрузки (если есть файл)
        if uploading_file_path and left_serial:
            # Проверяем наличие серийного номера (с логированием)
            found = check_serial_in_uploading_file(
                uploading_file_path, left_serial, log_func
            )
            if not found:
                # Попытка 1: добавить ведущий ноль
                serial_with_zero = check_serial_with_leading_zero(left_serial)
                if serial_with_zero != left_serial:
                    found = check_serial_in_uploading_file(
                        uploading_file_path, serial_with_zero, log_func
                    )
            if not found:
                # Попытка 2: изменить регистр на противоположный
                serial_swapped = swap_case_serial(left_serial)
                if serial_swapped != left_serial:
                    found = check_serial_in_uploading_file(
                        uploading_file_path, serial_swapped, log_func
                    )
            status = "Найден" if found else "Не найден"
            ws.cell(row=row_num, column=5, value=status)
            if not found:
                ws.cell(row=row_num, column=5).font = Font(
                    color="FF0000", bold=True
                )

        if right_rows:
            right_name, right_serial = right_row
            ws.cell(row=row_num, column=7, value=i)
            ws.cell(row=row_num, column=8, value=right_name)
            ws.cell(row=row_num, column=9, value=right_serial)

            # RMA правый (столбец J)
            if rma_dict and right_serial:
                rma_entry = rma_dict.get(right_serial)
                if rma_entry:
                    val, headers = rma_entry
                    cell_text = f"{val} ({headers})" if headers else val
                    ws.cell(row=row_num, column=10, value=cell_text)
                else:
                    ws.cell(row=row_num, column=10, value="")

    # Выделение несовпадений (индексы скорректированы)
    if right_rows:
        red_bold_font = Font(color="FF0000", bold=True)
        for row in range(2, ws.max_row + 1):
            serial_left = ws.cell(row=row, column=3).value
            serial_right = ws.cell(row=row, column=9).value

            # Серийные номера
            if serial_left and serial_right and serial_left != serial_right:
                ws.cell(row=row, column=3).font = red_bold_font
                ws.cell(row=row, column=9).font = red_bold_font
            else:
                if serial_left and not serial_right:
                    ws.cell(row=row, column=3).font = red_bold_font
                elif serial_right and not serial_left:
                    ws.cell(row=row, column=9).font = red_bold_font

    # Автоширина столбцов (A-E, G-J)
    for col_idx in [1, 2, 3, 4, 5, 7, 8, 9, 10]:
        if not right_rows and col_idx > 5:
            continue
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                continue
        adjusted_width = min(max_len + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_path)


# ----------------------------------------------------------------------
# Обработка информации по RMA
def load_rma_data_for_id(
    config: dict, normalized_id: str, log_func: Optional[Callable] = None
) -> dict:
    """
    Загружает данные из файла RMA (rma_tab) только для строк,
    где ID в столбце A соответствует normalized_id.
    Возвращает словарь:
        {серийный_номер: (значение_из_I, 'заголовок_K / заголовок_L')}
    Заголовки столбцов читаются из строки 3 (столбцы 11 и 12).
    """
    if log_func is None:
        def log_func(msg, level='info'):
            pass

    rma_path = config.get("rma_tab")
    rma_sheet = config.get("rma_tab_sheet")
    if not rma_path or not Path(rma_path).exists():
        log_func(f"Файл RMA не найден: {rma_path}", "warning")
        return {}

    try:
        wb = safe_load_workbook(rma_path, data_only=True)
        if rma_sheet not in wb.sheetnames:
            log_func(f"Лист '{rma_sheet}' не найден в файле RMA", "error")
            wb.close()
            return {}
        ws = wb[rma_sheet]

        # Читаем заголовки из строки 3 для столбцов K(11) и L(12)
        header_k = format_cell_value(ws.cell(row=3, column=11).value) or "K"
        header_l = format_cell_value(ws.cell(row=3, column=12).value) or "L"
        log_func(f"Заголовки RMA: K='{header_k}', L='{header_l}'", "debug")

        # Промежуточный словарь:
        # serial -> [значение_I, список_уникальных_заголовков]
        raw_dict = {}
        start_row = 4
        max_row = ws.max_row

        for row_idx in range(start_row, max_row + 1):
            cell_id = ws.cell(row=row_idx, column=1).value
            if cell_id is None:
                continue
            id_str = str(cell_id).strip()
            digits = re.sub(r'\D', '', id_str)
            if not digits:
                continue
            cell_norm = digits.zfill(4)
            if cell_norm != normalized_id:
                continue

            value_i = ws.cell(row=row_idx, column=9).value
            if value_i is None:
                continue
            val_str = format_cell_value(value_i)

            serial_k = ws.cell(row=row_idx, column=11).value
            serial_l = ws.cell(row=row_idx, column=12).value
            serial_k_str = format_cell_value(serial_k)
            serial_l_str = format_cell_value(serial_l)

            if serial_k_str:
                entry = raw_dict.setdefault(serial_k_str, [val_str, []])
                if header_k not in entry[1]:
                    entry[1].append(header_k)
            if serial_l_str:
                entry = raw_dict.setdefault(serial_l_str, [val_str, []])
                if header_l not in entry[1]:
                    entry[1].append(header_l)

        wb.close()

        # Преобразуем в финальный вид: {serial: (value, 'h1 / h2')}
        rma_dict = {
            k: (v[0], " / ".join(v[1]))
            for k, v in raw_dict.items()
        }
        log_func(
            f"Загружено {len(rma_dict)} записей из RMA для ID {normalized_id}",
            "info"
        )
        return rma_dict
    except Exception as e:
        log_func(f"Ошибка загрузки RMA: {e}", "error")
        return {}


# ----------------------------------------------------------------------
# Функции для проверки серийных номеров в папке folder_uploading_data
def find_newest_uploading_file(
    uploading_folder: Path, normalized_id: str
) -> Optional[Path]:
    """
    Находит самый новый файл xlsx в папке uploading_folder,
    имя которого начинается с числа, соответствующего normalized_id (без
    ведущих нулей).
    """
    if not uploading_folder.exists():
        return None
    id_raw = str(int(normalized_id))  # Убираем ведущие нули
    candidates = []
    for file_path in uploading_folder.glob("*.xlsx"):
        first_part = file_path.stem.split()[0] if file_path.stem else ""
        digits = re.sub(r'\D', '', first_part)
        if digits and int(digits) == int(id_raw):
            candidates.append(file_path)
    if not candidates:
        return None
    # Возвращаем самый новый по дате модификации
    return max(candidates, key=lambda p: p.stat().st_mtime)


def check_serial_in_uploading_file(
    uploading_file: Path, serial: str, log_func=None
) -> bool:
    """
    Проверяет, встречается ли серийный номер в столбцах D, E, F, G (4-7)
    в первых 500 строках файла uploading_file.
    Возвращает True, если найден хотя бы раз.
    """
    if not uploading_file or not uploading_file.exists():
        if log_func:
            log_func(
                f"Файл для проверки загрузки не существует: {uploading_file}",
                "warning"
            )
        return False
    try:
        wb = safe_load_workbook(uploading_file, data_only=True, read_only=True)
        ws = wb.active
        found = False
        # Проверяем строки 1..500, столбцы 4..7 (D,E,F,G)
        for row_idx, row in enumerate(ws.iter_rows(
            min_row=1, max_row=500, min_col=4, max_col=7, values_only=True
        ), start=1):
            for cell_value in row:
                if cell_value is not None:
                    cell_str = format_cell_value(cell_value)
                    if serial == cell_str:
                        found = True
                        break
            if found:
                break
        wb.close()
        if log_func and found:
            log_func(
                f"SN {serial} найден в файле {uploading_file.name}", "debug"
            )
        return found
    except Exception as e:
        if log_func:
            log_func(
                f"Ошибка при проверке файла {uploading_file}: {e}", "error"
            )
        return False


def check_serial_with_leading_zero(serial: str) -> str:
    """
    Добавляет один ведущий ноль к строке, если она состоит из цифр.
    Иначе возвращает исходную строку (без изменений).
    """
    if serial and serial.isdigit():
        return '0' + serial
    return serial


def swap_case_serial(serial: str) -> str:
    """Меняет регистр всех букв в строке на противоположный."""
    if not serial:
        return serial
    return ''.join(
        ch.lower()
        if ch.isupper()
        else ch.upper()
        if ch.islower()
        else ch
        for ch in serial
    )
