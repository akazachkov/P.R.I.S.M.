# /core/elements/report_collector.py

import re
from datetime import datetime
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


class ReportCollector:
    """Сбор данных из множества Excel-файлов в единый отчёт."""

    # Цвет заливки для ячеек с датой, взятой из свойств файла
    YELLOW_FILL = PatternFill(
        start_color="FFFF99", end_color="FFFF99", fill_type="solid"
    )

    def __init__(self, config: dict, year: str = None, log_func=None):
        self.log = log_func if log_func else print
        self.folder_pnr = Path(config["folder_pnr"])
        self.folder_pmo = Path(config["folder_pmo"])
        self.column_names = config["pnr_pmo_column_names"]
        self.rows = []

        if year:
            self.year = year
            target_key = f"rrp_data_{year}"
            if target_key not in config:
                raise ValueError(f"Ключ '{target_key}' не найден в конфиге")
        else:
            self.year = year
            # Поиск ключа вида "rrp_data_ГГГГ"
            target_key = None
            for key in config:
                if re.match(r"rrp_data_\d{4}$", key):
                    target_key = key
                    break
            if not target_key:
                raise ValueError(
                    "В конфиге не найден ключ с именем 'rrp_data_ГГГГ'"
                )
            self.year = target_key.split("_")[-1]

        self.target_dir = Path(config[target_key])
        self.target_file = self.target_dir / "Свод отчётов PMO, PNR.xlsx"
        # Сбор всех строк отчёта в памяти в кортеж
        self.rows = []

    def run(self, mode: str = "append"):
        """
        Запуск сбора данных.
        :param mode: 'overwrite' - перезаписать файл
                     'append' - дополнить файл с проверкой дубликатов
        :return: dict с ключами 'added' и 'total' (при append) или None
        """
        self.log("Начинаем сбор данных...")
        # Определяем дату последней записи в отчёте
        if mode == "append" and self.target_file.exists():
            self.last_date = self._get_last_report_date()
            if self.last_date:
                self.log(
                    "Последняя дата в отчёте: "
                    f"{self.last_date.strftime('%Y.%m.%d')}")
        else:
            self.last_date = None

        self._collect_files(self.folder_pnr)
        self._collect_files(self.folder_pmo)

        if mode == "overwrite":
            self._save_overwrite()
            return None
        elif mode == "append":
            return self._save_append()
        else:
            raise ValueError("mode должен быть 'overwrite' или 'append'")

    def _collect_files(self, folder: Path):
        """Обработка всех .xlsx файлов в корне указанной папки."""
        if not folder.exists():
            self.log(f"Внимание: папка {folder} не существует, пропускаем.")
            return

        for file_path in folder.glob("*.xlsx"):
            if not file_path.is_file():
                continue
            # Пропускаем файлы, которые не новее последней даты в отчёте
            if self.last_date:
                file_mtime = file_path.stat().st_mtime
                if file_mtime <= self.last_date.timestamp():
                    continue
            self.log(f"  Обработка файла: {file_path.name}")
            try:
                self._process_file(file_path)
            except Exception as e:
                self.log(f"  Ошибка при обработке {file_path.name}: {e}")

    def _process_file(self, file_path: Path):
        """Чтение одного Excel-файла и добавление его строк в self.rows."""
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Определяем дату для файла
        date_value, is_fallback = self._extract_date(file_path)
        # Формат: ГГГГ.ММ.ДД
        date_str = (
            f"{date_value.year}.{date_value.month:02d}.{date_value.day:02d}"
        )

        # Считываем заголовки (первая строка)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Определяем индексы нужных столбцов (по именам из конфига)
        col_indices = {}
        for col_name in self.column_names:
            try:
                idx = headers.index(col_name)
                col_indices[col_name] = idx
            except ValueError:
                # Столбец не найден — оставляем пустым
                pass

        # Если ни одного столбца не найдено — всё равно обрабатываем строки
        # (будут заполнены только Дата и Имя файла)

        # Обрабатываем строки данных (начиная со второй)
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Первые два столбца: Дата и Имя файла
            row_data = [date_str, file_path.stem]

            # Добавляем данные из указанных столбцов (если найдены)
            for col_name in self.column_names:
                if col_name in col_indices:
                    val = row[col_indices[col_name]]
                    row_data.append(val)
                else:
                    row_data.append(None)

            self.rows.append((row_data, is_fallback, date_value))

        wb.close()

    def _extract_date(self, file_path: Path) -> tuple[datetime, bool]:
        """
        Извлекает дату из имени файла.
        :return: (datetime, is_fallback) — где is_fallback = True, если взята
        дата изменения файла
        """
        stem = file_path.stem  # имя без расширения
        # Шаблон: дата в формате дд.мм или дд.мм-дд.мм в конце имени перед
        # расширением
        pattern = r"(\d{2}\.\d{2})(?:-(\d{2}\.\d{2}))?$"
        match = re.search(pattern, stem)

        # Дополнительно проверим, нет ли других чисел с точками в имени.
        # Если в имени есть точки, кроме найденной в конце, считаем, что
        # извлечь не удалось.
        if match:
            date_part = match.group(1)
            # Проверим, нет ли в остальной части имени (до совпадения) других
            # чисел с точками
            prefix = stem[: match.start()]
            if re.search(r"\d+\.\d+", prefix):
                # Есть другие числа с точками — fallback
                return self._get_file_mtime(file_path), True

            # Попробуем преобразовать в дату
            try:
                day, month = map(int, date_part.split("."))
                return datetime(int(self.year), month, day), False
            except ValueError:
                return self._get_file_mtime(file_path), True
        else:
            # Не найдено в конце — fallback
            return self._get_file_mtime(file_path), True

    def _get_file_mtime(self, file_path: Path) -> datetime:
        """Возвращает дату последнего изменения файла."""
        timestamp = file_path.stat().st_mtime
        return datetime.fromtimestamp(timestamp)

    def _get_last_report_date(self) -> Optional[datetime]:
        """
        Возвращает дату из последней строки столбца A существующего файла
        отчёта.
        """
        if not self.target_file.exists():
            return None
        try:
            wb = load_workbook(
                self.target_file, data_only=True, read_only=True
            )
            ws = wb.active
            last_row = ws.max_row
            if last_row < 2:  # только заголовок
                wb.close()
                return None
            date_str = ws.cell(row=last_row, column=1).value
            wb.close()
            if date_str is None:
                return None
            # Формат даты в файле: ГГГГ.ММ.ДД
            return datetime.strptime(str(date_str), "%Y.%m.%d")
        except Exception as e:
            self.log(f"Не удалось прочитать последнюю дату из отчёта: {e}")
            return None

    def _save_overwrite(self):
        """
        Создаёт новый файл (или перезаписывает существующий) со всеми
        данными.
        """
        self.target_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Сбор данных"

        # Заголовки
        headers = ["Дата", "Имя файла"] + self.column_names
        ws.append(headers)

        # Данные
        for row_data, need_yellow, _ in self.rows:
            ws.append(row_data)
            if need_yellow:
                # Закрашиваем ячейку с датой (первый столбец текущей строки)
                last_row = ws.max_row
                cell = ws.cell(row=last_row, column=1)
                cell.fill = self.YELLOW_FILL

        # Автоширина для удобства (опционально)
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        wb.save(self.target_file)
        self.log(f"Файл перезаписан: {self.target_file}")

    def _save_append(self):
        """
        Дополняет существующий файл новыми данными, избегая дубликатов.
        Возвращает dict с ключами 'added' и 'total'.
        """
        if not self.target_file.exists():
            # Если файла нет — создаём как при перезаписи
            self._save_overwrite()
            return {'added': 0, 'total': len(self.rows)}

        # Используем self.last_date, полученное в run
        if self.last_date:
            filtered_rows = []
            for row_data, need_yellow, date_value in self.rows:
                if date_value > self.last_date:
                    filtered_rows.append((row_data, need_yellow))
            self.log(
                "Отфильтровано строк: "
                f"{len(self.rows) - len(filtered_rows)} из {len(self.rows)}"
            )
        else:
            filtered_rows = [
                (row_data, need_yellow) for row_data,
                need_yellow, _ in self.rows
            ]

        wb = load_workbook(self.target_file)
        ws = wb.active

        # Собираем существующие пары (Дата, Имя файла)
        existing_pairs = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                existing_pairs.add((row[0], row[1]))

        # Добавляем только новые строки
        added = 0
        for row_data, need_yellow in filtered_rows:
            pair = (row_data[0], row_data[1])
            if pair not in existing_pairs:
                ws.append(row_data)
                if need_yellow:
                    last_row = ws.max_row
                    cell = ws.cell(row=last_row, column=1)
                    cell.fill = self.YELLOW_FILL
                added += 1

        wb.save(self.target_file)
        total = ws.max_row - 1
        self.log(f"Новых строк: {added}. Всего строк в файле: {total}")
        return {'added': added, 'total': total}
